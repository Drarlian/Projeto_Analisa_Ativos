from manipula_planilha_google import *
from manipula_planilha_excel import *


def analisar_pvp(tipo_ativo: str) -> None:
    """
    Pega a lista dos ativos informados presentes na planilha e analisa o valor do P/PV, de acordo com a analise altera
    o fundo do valor na planilha. (verde = bom | vermelho = ruim)
    :param tipo_ativo: Determina o tipo de ativo -> acoes | fiis
    :return: None
    """
    indicador_positivo = 1.05

    if tipo_ativo == 'acoes':
        lista_ativos = pegar_dados_planilha('Página1!A2:G')
        posicao_do_elemento_pvp = 4
        colunas = (4, 5)  # -> Começa na coluna 4 e termina na coluna 5.
        # -> (A primeira coluna não é alterada, apenas as colunas seguintes da primeira)
    elif tipo_ativo == 'fiis':
        lista_ativos = pegar_dados_planilha('Página1!I2:O')
        posicao_do_elemento_pvp = 3
        colunas = (11, 12)  # -> Começa na coluna 11 e termina na coluna 12.
        # -> (A primeira coluna não é alterada, apenas as colunas seguintes da primeira)
    else:
        return None

    if lista_ativos[0][0] == 'Ativo':  # -> Nunca deve acontecer.
        lista_ativos.pop(0)

    for c in range(len(lista_ativos)):
        if float(lista_ativos[c][posicao_do_elemento_pvp].replace(',', '.')) <= indicador_positivo:
            cell = criar_celula(base=False, cores=(0, 1, 0))
        else:
            cell = criar_celula(base=False, cores=(1, 0, 0))

        atualizar_formatacao_planilha(base=False,
                                      request=criar_request(cell, base=False, linhas=(c+1, c+2), colunas=colunas))


def criar_celula(base: bool = True, alinhamento: str = 'LEFT', cores: tuple = (1, 1, 1)):
    if base:
        cell_format = {
            'horizontalAlignment': 'LEFT',
            'backgroundColor': {
                'red': 1,
                'green': 1,
                'blue': 1
            }
        }
    else:
        cell_format = {
            'horizontalAlignment': alinhamento,
            'backgroundColor': {
                'red': cores[0],
                'green': cores[1],
                'blue': cores[2]
            }
        }

    return cell_format


def criar_request(cell_format, base: bool = True, linhas: tuple = (0, 99), colunas: tuple = (0, 99)):
    if base:
        request = {
            'repeatCell': {
                'range': {
                    'sheetId': 0,
                    'startRowIndex': 0,
                    'endRowIndex': 99,
                    'startColumnIndex': 0,
                    'endColumnIndex': 99
                },
                'cell': {
                    'userEnteredFormat': cell_format
                },
                'fields': 'userEnteredFormat.horizontalAlignment,userEnteredFormat.backgroundColor'
            }
        }
    else:
        request = {
            'repeatCell': {
                'range': {
                    'sheetId': 0,
                    'startRowIndex': linhas[0],
                    'endRowIndex': linhas[1],
                    'startColumnIndex': colunas[0],
                    'endColumnIndex': colunas[1]
                },
                'cell': {
                    'userEnteredFormat': cell_format
                },
                'fields': 'userEnteredFormat.horizontalAlignment,userEnteredFormat.backgroundColor'
            }
        }

    return request


def descobre_range(lista: list) -> str:
    """
    Descobre o range que vai ser usado pela lista recebida.
    """
    import string
    letras = list(string.ascii_uppercase)

    return f'Página1!A2:{letras[len(lista[0]) - 1]}'

def analisar_pvp_excel(tipo_ativo: str, lista_ativos: list,  todos: bool = False) -> None:
    """
    Recebe uma lista com os ativos que serão atualizados.
    :param tipo_ativo: Tipo de ativo que será analisado | Possibilidades: 'fiis' ou 'acoes'
    :param lista_ativos: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    from openpyxl.styles import PatternFill

    if tipo_ativo == 'acoes':
        intervalo = 'A2:F'
        posicao_do_elemento_pvp = 4
    elif tipo_ativo == 'fiis':
        intervalo = 'I2:N'
        posicao_do_elemento_pvp = 3
    else:
        return None

    intervalo = intervalo + descobrir_ultima_linha_planilha()  # Descobrindo o número da última linha preenchida.

    if todos:
        ativos = pegar_dados_intervalo_planilha(intervalo)
        lista_ativos = [elemento[0] for elemento in ativos]

    indicador_positivo = 1.05

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    cont = 0
    for celula in aba_ativa[intervalo]:
        if celula[0].value is None:
            break
        else:
            if celula[0].value == lista_ativos[cont]:
                # Se o indicador for negativo:
                if celula[posicao_do_elemento_pvp].value >= indicador_positivo:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                       fill_type='solid')
                # Se o indicador for positivo:
                else:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='008000', end_color=None,
                                                                       fill_type='solid')

                if lista_ativos[cont] == lista_ativos[-1]:
                    break
                else:
                    cont += 1

    planilha.save('Arquivo.xlsx')
    planilha.close()


def descobrir_ultima_linha_planilha() -> str:
    """
    Descobre o número da ultima linha preenchida na planilha.
    :return: Retorna o número da ultima linha como uma string.
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    ultima_linha = aba_ativa['A'][-1].row

    planilha.close()

    return str(ultima_linha)


if __name__ == '__main__':
    # analisar_pvp('acoes')
    # analisar_pvp('fiis')

    analisar_pvp_excel('fiis', lista_ativos=None, todos=True)
    analisar_pvp_excel('acoes', lista_ativos=None, todos=True)
