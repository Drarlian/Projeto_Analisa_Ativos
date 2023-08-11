from manipula_planilha_excel import *


def analisar_pvp_excel(tipo_ativo: str, lista_ativos: list,  todos: bool = False) -> None:
    """
    Recebe uma lista com os ativos que serão atualizados.
    :param tipo_ativo: Tipo de ativo que será analisado | Possibilidades: 'fiis' ou 'acoes'
    :param lista_ativos: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    from openpyxl.styles import PatternFill

    if tipo_ativo == 'acoes':
        intervalo = 'A2:F'
        posicao_do_elemento_pvp = 4
    elif tipo_ativo == 'fiis':
        intervalo = 'I2:N'
        posicao_do_elemento_pvp = 3
    else:
        return None

    intervalo = intervalo + descobrir_ultima_linha_planilha_excel(intervalo[0])  # Descobrindo o número da última linha preenchida.

    if todos:
        ativos = pegar_dados_intervalo_planilha(intervalo)
        lista_ativos = [elemento[0] for elemento in ativos]

    indicador_positivo = 1.05

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    cont = 0
    for celula in aba_ativa[intervalo]:
        if celula[0].value is None:
            break
        else:
            if celula[0].value == lista_ativos[cont]:
                # Se o indicador for negativo:
                if float(celula[posicao_do_elemento_pvp].value.replace(',', '.')) >= indicador_positivo:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                       fill_type='solid')
                # Se o indicador for positivo:
                else:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='008000', end_color=None,
                                                                       fill_type='solid')

                if lista_ativos[cont] == lista_ativos[-1]:
                    break
                else:
                    cont += 1

    planilha.save('Arquivo.xlsx')
    planilha.close()


def descobrir_ultima_linha_planilha_excel(coluna: str) -> str:
    """
    Descobre o número da ultima linha preenchida na planilha da coluna informada.
    :return: Retorna o número da ultima linha como uma string.
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    ultima_linha = aba_ativa[coluna][-1].row

    planilha.close()

    return str(ultima_linha)


if __name__ == '__main__':
    # analisar_pvp_excel('fiis', ['RBRY11', 'VGIR11'])
    # analisar_pvp_excel('fiis', ['DEVA11', 'HGLG11'])
    analisar_pvp_excel('fiis', lista_ativos=None, todos=True)

    # analisar_pvp_excel('acoes', ['TAEE4', 'BBDC4'])
    # analisar_pvp_excel('acoes', ['JBSS3', 'EKTR3'])
    analisar_pvp_excel('acoes', lista_ativos=None, todos=True)
