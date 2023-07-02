from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def iniciar_planilha():
    """
    Inicia o arquivo Excel contendo os dados.
    :return: Retorna a planilha.
    """
    try:
        open('Arquivo.xlsx', 'a').close()
        planilha = load_workbook('Arquivo.xlsx')
    except PermissionError:
        print('O arquivo já está aberto, feche o mesmo antes de prosseguir.')
        exit()
    except:
        print('Arquivo não encontrado.')
        exit()
    else:
        return planilha


def pegar_dados_intervalo_planilha(intervalo: str) -> list:
    """
    Retorna os valores presentes no intervalo informado.
    Os valores são retornados dentro de uma lista.
    A lista retornada contém listas para cada linha do intervalo informado.
    Ex: [['Pessoa1', 46, 2500, 'Jogador', '987654321'], ['Pessoa2', 22, 8000, 'Streamer', '768948302']]
    :param intervalo: Intervalo da planilha.
    :return: Retorna uma lista contendo os valores.
    """

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        valores: list = []
        valores_linha: list = []

        # Adicionei os if's para impedir que dados vazios sejam obtidos.
        for celula in aba_ativa[intervalo]:
            for elemento in celula:
                if elemento.value is not None:
                    valores_linha.append(elemento.value)
                else:
                    break  # -> Talvez eu possa tirar esse else e deixar ele pegar uma linha onde um elemento seja None.
            if len(valores_linha) > 0:
                valores.append(valores_linha.copy())
                valores_linha.clear()
            else:
                break  # -> Para a procura se achar algum registro vazio.
    except:
        planilha.close()
        print('Error!')
    else:
        planilha.close()
        return valores


def atualizar_dados_intervalo_planilha(valores_adicionar: list, intervalo: str) -> None:
    """
    Atualiza os dados presentes no intervalo informado.
    Caso o intervalo esteja vazio, um erro de Index é gerado.
    :param valores_adicionar: Valores que vão substituir os dados presentes no intervalo.
    :param intervalo: Intervalo que será substituído.
    :return: None
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            cont = 0
            for elemento in celula:
                if elemento.value is None:  # -> Gerando o erro de função específica.
                    raise IndexError
                elemento.value = valores_adicionar[cont]
                cont += 1
    except IndexError:
        print(f'Uma ou mais células não possuem um valor para atualizar.')
    except:
        print('Error')
    else:
        planilha.save('Arquivo.xlsx')
    finally:
        planilha.close()


def adicionar_dados_fim_coluna(valores_adicionar: list, coluna: str) -> None:
    """
    Adiciona os valores na primeira linha disponível da coluna informada.
    :param valores_adicionar: Valores a serem adicionados.
    :param coluna: Coluna onde os valores serão adicionados.
    :return: None
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[coluna]:
            if celula.value is None:
                coluna_numero = celula.column
                for elemento in range(len(valores_adicionar)):
                    coluna = get_column_letter(coluna_numero)
                    aba_ativa[f'{coluna}{celula.row}'] = valores_adicionar[elemento]
                    coluna_numero += 1
                break
    except:
        print('Error')
    else:
        planilha.save('Arquivo.xlsx')
    finally:
        planilha.close()


def adicionar_dados_intervalo_planilha(valores_adicionar: list, intervalo: str) -> None:
    """
    Adiciona os dados informados no intervalo especificado da planilha.
    Caso já existam dados nesse intervalo, gera um erro de Index.
    :param valores_adicionar: Dados a serem adicionados.
    :param intervalo: Intervalo da planilha.
    :return: None
    """

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            cont = 0
            for elemento in celula:
                if elemento.value is not None:  # -> Gerando o erro de função específica.
                    raise IndexError
                elemento.value = valores_adicionar[cont]
                cont += 1
    except IndexError:
        print(f'Uma ou mais células já possuem um valor.')
    except:
        print('Error')
    else:
        planilha.save('Arquivo.xlsx')
    finally:
        planilha.close()


def remover_dados_intervalo_planilha(intervalo: str) -> None:
    """
    Apaga os dados presentes no intervalo informado.
    Caso o intervalo esteja vazio, um erro de Index é gerado.
    :param intervalo: Intervalo que será apagado.
    :return: None
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            for elemento in celula:
                if elemento.value is None:  # -> Gerando o erro de função específica.
                    raise IndexError
                elemento.value = None
    except IndexError:
        print('Uma ou mais células já estão vazias um valor.')
    except:
        print('Error')
    else:
        planilha.save('Arquivo.xlsx')
    finally:
        planilha.close()


def atualizar_cor_intervalo_planilha(intervalo: str, cor: str = 'FFFFFFFF') -> None:
    """
    Atualiza a cor de fundo do intervalo informado.
    Caso seja informado apenas uma célula, apenas a cor dessa célula será atualizada.
    :param intervalo: Intervalo ou Célula a ser alterado.
    :param cor: Cor de fundo desejada no formato Hexadecimal, por padrão é branco.
    :return: None
    """
    from openpyxl.styles import PatternFill

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        if ':' in intervalo:
            for celula in aba_ativa[intervalo]:
                celula[0].fill = PatternFill(start_color=cor, end_color=None, fill_type='solid')
        else:
            # Aplica a formatação de preenchimento à célula A1 com cor sólida
            aba_ativa[intervalo].fill = PatternFill(start_color=cor, end_color=None, fill_type='solid')
    except:
        print('Error')
    else:
        planilha.save('Arquivo.xlsx')
    finally:
        planilha.close()


if __name__ == '__main__':
    print(pegar_dados_intervalo_planilha('A2:F99'))
