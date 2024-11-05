from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# import openpyxl


def iniciar_planilha():
    """
    Inicia o arquivo Excel contendo os dados.
    :return: Retorna a planilha.
    """
    try:
        open('PlanilhaExcel\\Arquivo.xlsx', 'a').close()
        planilha = load_workbook('PlanilhaExcel\\Arquivo.xlsx')
    except PermissionError:
        print('O arquivo já está aberto, feche o mesmo antes de prosseguir.')
        exit()
    except:
        print('Arquivo não encontrado.')
        exit()
    else:
        return planilha


def pegar_dados_intervalo_planilha(intervalo: str, ultima_linha: bool = False) -> list:
    """
    Retorna os valores presentes no intervalo informado.
    Os valores são retornados dentro de uma lista.
    A lista retornada contém listas para cada linha do intervalo informado.
    Ex: [['Pessoa1', 46, 2500, 'Jogador', '987654321'], ['Pessoa2', 22, 8000, 'Streamer', '768948302']]
    :param intervalo: Intervalo da planilha.
    :param ultima_linha: Define se deverá ser pego até a ultima linha do intervalo informado.
    :return: Retorna uma lista contendo os valores.
    """
    if ultima_linha:
        intervalo: str = intervalo + descobrir_linha_vazia_planilha_excel(intervalo[0])

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        valores: list = []
        valores_linha: list = []

        # Adicionei os if's para impedir que dados vazios sejam obtidos.
        for celula in aba_ativa[intervalo]:
            for elemento in celula:
                if elemento.value is not None:
                    valores_linha.append(elemento.value)
                else:
                    break  # -> Talvez eu possa tirar esse else e deixar ele pegar uma linha onde um elemento seja None.
            if len(valores_linha) > 0:
                valores.append(valores_linha.copy())
                valores_linha.clear()
            # else:
                # break  # -> Para a procura se achar algum registro vazio.
    except:
        planilha.close()
        print('Error - pegar_dados_intervalo_planilha()')
    else:
        planilha.close()
        return valores


def atualizar_dados_intervalo_planilha(valores_adicionar: list, intervalo: str) -> None:
    """
    Atualiza os dados presentes no intervalo informado.
    Caso o intervalo esteja vazio, um erro de Index é gerado.  <- FUNCIONALIDADE DESATIVADA
    :param valores_adicionar: Lista de listas contendo os valores que vão substituir os dados presentes no intervalo.
    :param intervalo: Intervalo que será substituído.
    :return: None
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for i, linha in enumerate(aba_ativa[intervalo]):  # -> A2, B2, C2, ...
            if i >= len(valores_adicionar):
                break
            for j, elemento in enumerate(linha):
                # if elemento.value is None:  # -> Gerando o erro de função específica.
                    # raise IndexError("Não existe valor na célula.")
                elemento.value = valores_adicionar[i][j]
    except IndexError:
        print(f'Error - atualizar_dados_intervalo_planilha() | Uma ou mais células não possuem um valor para atualizar.')
    except:
        print('Error - atualizar_dados_intervalo_planilha()')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def adicionar_dados_fim_coluna(valores_adicionar: list, coluna_inicial: str, coluna_final: str) -> None:
    """
    Adiciona todos os itens da lista informada nas linhas disponíveis na coluna informada.
    :param valores_adicionar: Lista de listas contendo os valores a serem adicionados.
    :param coluna_inicial: Primeira coluna onde os valores serão adicionados.
    :param coluna_final: Última coluna onde os valores serão adicionados.
    :return: None
    """
    try:
        ultima_linha: str = descobrir_linha_vazia_planilha_excel(coluna_inicial)
        comeco = int(ultima_linha) + 1
        fim = int(ultima_linha) + len(valores_adicionar)
        intervalo: str = f'{coluna_inicial}{comeco}:{coluna_final}{fim}'

        adicionar_dados_intervalo_planilha(valores_adicionar, intervalo)
    except:
        print('Error - adicionar_dados_fim_coluna()')


def adicionar_dados_intervalo_planilha(valores_adicionar: list, intervalo: str, ultima_linha: bool = False) -> None:
    """
    Adiciona os dados informados no intervalo especificado da planilha.
    Caso já existam dados nesse intervalo, gera um erro de Index. <- FUNCIONALIDADE DESATIVADA
    :param valores_adicionar: Lista de listas contendo os dados a serem adicionados.
    :param intervalo: Intervalo da planilha.
    :param ultima_linha: Define se deverá ser pego até a ultima linha do intervalo informado. 'A2:E'
    :return: None
    """
    if ultima_linha:
        intervalo: str = intervalo + descobrir_linha_vazia_planilha_excel(intervalo[0])

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for i, linha in enumerate(aba_ativa[intervalo]):  # -> A2, B2, C2, ...
            if i >= len(valores_adicionar):
                break
            for j, elemento in enumerate(linha):
                # if elemento.value is not None:  # -> Gerando o erro de função específica.
                    # raise IndexError("Já existe um valor na célula.")
                elemento.value = valores_adicionar[i][j]
    # except IndexError:
        # print(f'Error - adicionar_dados_intervalo_planilha() | Uma ou mais células já possuem um valor.')
    except:
        print('Error - adicionar_dados_intervalo_planilha()')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def remover_dados_intervalo_planilha(intervalo: str, ultima_linha: bool = False) -> None:
    """
    Apaga os dados presentes no intervalo informado.
    Caso o intervalo esteja vazio, um erro de Index é gerado. <- FUNCIONALIDADE DESATIVADA
    :param intervalo: Intervalo que será apagado.
    :param ultima_linha: Define se deverá ser pego até a ultima linha do intervalo informado. 'A2:E'
    :return: None
    """
    if ultima_linha:
        intervalo: str = intervalo + descobrir_linha_vazia_planilha_excel(intervalo[0])

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            for elemento in celula:
                # if elemento.value is None:  # -> Gerando o erro de função específica.
                    # raise IndexError
                elemento.value = None
    # except IndexError:
        # print('Error - remover_dados_intervalo_planilha() | Uma ou mais células já estão vazias.')
    except:
        print('Error - remover_dados_intervalo_planilha()')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def atualizar_cor_intervalo_planilha(intervalo: str, cor: str = 'FFFFFFFF') -> None:
    """
    Atualiza a cor de fundo do intervalo informado.
    Caso seja informado apenas uma célula, apenas a cor dessa célula será atualizada.
    :param intervalo: Intervalo ou Célula a ser alterado.
    :param cor: Cor de fundo desejada no formato Hexadecimal, por padrão é branco.
    :return: None
    """
    from openpyxl.styles import PatternFill

    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    try:
        if ':' in intervalo:
            for celula in aba_ativa[intervalo]:
                celula[0].fill = PatternFill(start_color=cor, end_color=None, fill_type='solid')
        else:
            # Aplica a formatação de preenchimento à célula A1 com cor sólida
            aba_ativa[intervalo].fill = PatternFill(start_color=cor, end_color=None, fill_type='solid')
    except:
        print('Error - atualizar_cor_intervalo_planilha()')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def descobrir_ultima_linha_planilha_excel() -> str:
    """
    Descobre o número da ultima linha preenchida na planilha.
    Independente da coluna sempre vai retornar a última linha preenchida na planilha. (De qualquer coluna)
    :return: Retorna o número da ultima linha como uma string.
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    ultima_linha: int = aba_ativa['A'][-1].row

    planilha.close()

    return str(ultima_linha)


def descobrir_linha_vazia_planilha_excel(coluna: str) -> str:
    """
    Descobre o número da ultima linha preenchida na planilha da coluna informada.
    Mesmo se houver linhas vazias entres linhas preenchidas, a última será pega.
    :return: Retorna o número da ultima linha como uma string.
    """
    planilha = iniciar_planilha()
    aba_ativa = planilha.active

    ultima_posicao: int = aba_ativa[coluna][-1].row

    ultima_linha: int = 1
    cont: int = 0

    for i, celula in enumerate(aba_ativa[f'{coluna}1:{coluna}{ultima_posicao}']):
        if celula[0].value is not None:
            ultima_linha: int = celula[0].row
        else:
            # Ignoro a quantidade de linhas que eu percebi que eu preciso passar para chegar na proxima linha NÃO vazia:
            if cont > 0:
                cont -= 1  # Diminuo o cont pois acabei de passar por uma linha que estou ignorando.
            else:
                # aba_ativa[coluna][i].row == celula[0].row  | Ambos são a mesma coisa.
                if aba_ativa[coluna][i].row < ultima_posicao:
                    # VERIFICANDO SE EXISTE ALGUMA LINHA PREENCHIDA DA POSIÇÃO ATUAL ATE A ULTIMA LINHA DA COLUNA:
                    quantidade_linha_faltante: int = ultima_posicao - celula[0].row
                    elemento: int = 0  # -> Variável utilizada para verificar se foi encontrada uma linha NÃO vazia.
                    # Loop baseado na quantidade de linhas faltantes, se faltam 7 linhas para acabar,
                    # o loop vai rodar 7 vezes.
                    for c in range(1, quantidade_linha_faltante+1):
                        valor = aba_ativa[coluna][i+c].value  # PEGANDO O VALOR PRESENTE NA LINHA *SEGUINTE*.
                        # ^ Para verificar em seguida se a mesma é uma linha vazia ou não.
                        if valor is not None:
                            elemento: int = 1  # -> Indicador de que encontrei uma linha NÃO vazia.
                            cont = c-1  # -> Quantidade de linhas que posso ignorar até a proxima linha NÃO vazia.
                            break
                            # ^ Se encontro alguma linha não vazia no meio do caminho, nem continuo olhando os próximos,
                            # pois já achei o que queria.
                    if elemento == 0:  # Se não encontrei nenhuma linha preenchida saio do loop e já sei a última linha.
                        break
                else:
                    break

    planilha.close()

    return str(ultima_linha)


if __name__ == '__main__':
    adicionar_dados_intervalo_planilha(['Pessoa54', 54, 2200, 'Professor', '81928348143'], 'A7:E7')
    print(pegar_dados_intervalo_planilha('A2:E5'))
    atualizar_dados_intervalo_planilha(['Pessoa1', 22, 8500, 'Streamer', '768948302'], 'A5:E5')
    print(pegar_dados_intervalo_planilha('A4:E4'))
    remover_dados_intervalo_planilha('A4:E4')
    print(pegar_dados_intervalo_planilha('A4:E4'))
    adicionar_dados_fim_coluna(['Pessoa24', 18, 500, 'Hipe', '768948302'], 'A')
    atualizar_cor_intervalo_planilha("D2:D6", 'FFFF0000')
    print(pegar_dados_intervalo_planilha('A2:F99'))
    print(pegar_dados_intervalo_planilha('A2:A', ultima_linha=True))
    lista = pegar_dados_intervalo_planilha('A2:A', ultima_linha=True)
    print(lista)
    pass
