import PlanilhaExcel.manipula_planilha_excel as manipula_excel


def analisar_ativos_excel(tipo_ativo: str, lista_ativos: list, todos: bool = False) -> None:
    """
    Função intermediaria responsável por chamar as funções de analise.
    :param tipo_ativo: Tipo de ativo que será analisado | Possibilidades: 'fiis' ou 'acoes'
    :param lista_ativos: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    if tipo_ativo not in ('acoes', 'fiis'):
        return None
    elif tipo_ativo == 'acoes':  # -> Analises únicas de acoes.
        analisar_pl_excel(lista_ativos, todos)

    # PROBLEMA COM OS FIIS EM ANALISA DY E VALORIZACAO
    analisar_pvp_excel(tipo_ativo, lista_ativos, todos)
    analisar_dy_excel(tipo_ativo, lista_ativos, todos)
    analisar_valorizacao_excel(tipo_ativo, lista_ativos, todos)


def analisar_pvp_excel(tipo_ativo: str, lista_ativos: list, todos: bool = False) -> None:
    """
    Recebe uma lista com os ativos que terão seu P/VP analisado.
    :param tipo_ativo: Tipo de ativo que será analisado | Possibilidades: 'fiis' ou 'acoes'
    :param lista_ativos: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    from openpyxl.styles import PatternFill

    if tipo_ativo == 'acoes':
        intervalo: str = 'A2:F'
        posicao_do_elemento_pvp: int = 4
    elif tipo_ativo == 'fiis':
        intervalo: str = 'I2:N'
        posicao_do_elemento_pvp: int = 3
    else:
        return None

    # Descobrindo o número da última linha preenchida:
    ultima_linha: str = manipula_excel.descobrir_linha_vazia_planilha_excel(intervalo[0])
    intervalo: str = intervalo + ultima_linha

    if todos:
        ativos: list = manipula_excel.pegar_dados_intervalo_planilha(intervalo)
        lista_ativos: list = [elemento[0] for elemento in ativos]

    valor_maximo_aceitavel: float = 1.05

    planilha = manipula_excel.iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            # Se o ativo atual da planilha estiver dentro da lista_ativos, ele analisa o ativo:
            if celula[0].value in lista_ativos:
                try:
                    if '-' not in celula[posicao_do_elemento_pvp].value:
                        valor_celula: float | None = float(celula[posicao_do_elemento_pvp].value.replace(',', '.'))
                    else:
                        valor_celula = None
                except AttributeError:
                    valor_celula: float | None = float(celula[posicao_do_elemento_pvp].value)

                # Caso não venha o valor do P/VP:
                if valor_celula is None:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                       fill_type='solid')
                # Se o valor do P/VP estiver acima do valor máximo aceitável: (vermelho)
                elif valor_celula >= valor_maximo_aceitavel:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                       fill_type='solid')
                # Se o valor do P/VP estiver abaixo do valor máximo aceitável: (verde)
                else:
                    celula[posicao_do_elemento_pvp].fill = PatternFill(start_color='008000', end_color=None,
                                                                       fill_type='solid')

                lista_ativos.remove(celula[0].value)  # -> Remove o ativo que foi analisado da lista_ativos.
                if len(lista_ativos) == 0:  # -> Se não existem mais ativos para serem analisados, encerra o loop.
                    break
    except Exception as e:
        print(f'Error! - 1 - {e}')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def analisar_pl_excel(lista_acoes: list, todos: bool = False) -> None:
    """
    Recebe uma lista com as AÇÕES que terão seu P/L analisados.
    :param lista_acoes: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    from openpyxl.styles import PatternFill

    intervalo: str = 'A2:F'
    posicao_do_elemento_pl: int = 3

    # Descobrindo o número da última linha preenchida:
    ultima_linha: str = manipula_excel.descobrir_linha_vazia_planilha_excel(intervalo[0])
    intervalo: str = intervalo + ultima_linha

    if todos:
        outro_intervalo: str = 'A2:A' + ultima_linha  # -> Pode ser assim pois sempre vai ser uma AÇÃO.
        ativos: list = manipula_excel.pegar_dados_intervalo_planilha(
            outro_intervalo)  # -> Pegando apenas o nome dos ativos.
        lista_acoes: list = [elemento[0] for elemento in ativos]

    valor_maximo_aceitavel: int = 10

    planilha = manipula_excel.iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            # Se o ativo atual da planilha estiver dentro da lista_ativos, ele analisa o ativo:
            if celula[0].value in lista_acoes:
                try:
                    valor_celula: float = float(celula[posicao_do_elemento_pl].value.replace(',', '.'))
                except AttributeError:
                    valor_celula: float = float(celula[posicao_do_elemento_pl].value)
                # Se o valor do P/L estiver acima do valor máximo aceitável: (vermelho)
                if valor_celula >= valor_maximo_aceitavel:
                    celula[posicao_do_elemento_pl].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                      fill_type='solid')
                # Se o valor do P/L for negativo: (amarelo)
                elif valor_celula < 0:
                    celula[posicao_do_elemento_pl].fill = PatternFill(start_color='FFFF00', end_color=None,
                                                                      fill_type='solid')
                # Se o valor do P/L estiver abaixo do valor máximo aceitável: (verde)
                else:
                    celula[posicao_do_elemento_pl].fill = PatternFill(start_color='008000', end_color=None,
                                                                      fill_type='solid')

                lista_acoes.remove(celula[0].value)  # -> Remove o ativo que foi analisado da lista_ativos.
                if len(lista_acoes) == 0:  # -> Se não existem mais ativos para serem analisados, encerra o loop.
                    break
    except Exception as e:
        print(f'Error! - 2 - {e}')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def analisar_dy_excel(tipo_ativo: str, lista_ativos: list, todos: bool = False) -> None:
    """
    Recebe uma lista com os ativos que terão seu Dividend Yield analisado.
    :param tipo_ativo: Tipo de ativo que será analisado | Possibilidades: 'fiis' ou 'acoes'
    :param lista_ativos: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    from openpyxl.styles import PatternFill

    if tipo_ativo == 'acoes':
        intervalo: str = 'A2:F'
        posicao_do_elemento_dy: int = 5
    elif tipo_ativo == 'fiis':
        intervalo: str = 'I2:N'
        posicao_do_elemento_dy: int = 2
    else:
        return None

    # Descobrindo o número da última linha preenchida:
    ultima_linha: str = manipula_excel.descobrir_linha_vazia_planilha_excel(intervalo[0])
    intervalo: str = intervalo + ultima_linha

    if todos:
        ativos: list = manipula_excel.pegar_dados_intervalo_planilha(intervalo)
        lista_ativos: list = [elemento[0] for elemento in ativos]

    valor_minimo_aceitavel: int = 8

    planilha = manipula_excel.iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            # Se o ativo atual da planilha estiver dentro da lista_ativos, ele analisa o ativo:
            if celula[0].value in lista_ativos:
                try:  # Caso o valor venha como uma string.
                    valor_celula: str = celula[posicao_do_elemento_dy].value.replace(',', '.')  # -> Alterando a ,
                    if '%' in valor_celula:
                        valor_celula: float = float(valor_celula.replace('%', ''))  # -> Alterando a %
                    else:
                        valor_celula: float = float(valor_celula) * 100  # -> Trazendo o valor de volta ao normal.
                except AttributeError:  # Caso o valor já venha como float.
                    valor_celula: float = celula[posicao_do_elemento_dy].value
                    valor_celula: float = float(valor_celula) * 100
                # Se o valor do DY estiver acima do valor minimo aceitável: (verde)
                if valor_celula >= valor_minimo_aceitavel:
                    celula[posicao_do_elemento_dy].fill = PatternFill(start_color='008000', end_color=None,
                                                                      fill_type='solid')
                # Se o valor do DY estiver abaixo do valor minimo aceitável: (vermelho)
                else:
                    celula[posicao_do_elemento_dy].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                      fill_type='solid')

                lista_ativos.remove(celula[0].value)  # -> Remove o ativo que foi analisado da lista_ativos.
                if len(lista_ativos) == 0:  # -> Se não existem mais ativos para serem analisados, encerra o loop.
                    break
    except Exception as e:
        print(f'Error! - 3 - {e}')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


def analisar_valorizacao_excel(tipo_ativo: str, lista_ativos: list, todos: bool = False) -> None:
    """
    Recebe uma lista com os ativos que terão sua Valorização analisada.
    :param tipo_ativo: Tipo de ativo que será analisado | Possibilidades: 'fiis' ou 'acoes'
    :param lista_ativos: Lista com ativos | Formato: ['ativo1', 'ativo2']
    :param todos: Define se TODOS os ativos do tipo informado serão atualizados.
    :return: None
    """
    from openpyxl.styles import PatternFill

    if tipo_ativo == 'acoes':
        intervalo: str = 'A2:F'
        posicao_do_elemento_valorizacao: int = 2
    elif tipo_ativo == 'fiis':
        intervalo: str = 'I2:N'
        posicao_do_elemento_valorizacao: int = 5
    else:
        return None

    # Descobrindo o número da última linha preenchida:
    ultima_linha: str = manipula_excel.descobrir_linha_vazia_planilha_excel(intervalo[0])
    intervalo: str = intervalo + ultima_linha

    if todos:
        ativos: list = manipula_excel.pegar_dados_intervalo_planilha(intervalo)
        lista_ativos: list = [elemento[0] for elemento in ativos]

    valor_minimo_aceitavel: int = 0

    planilha = manipula_excel.iniciar_planilha()
    aba_ativa = planilha.active

    try:
        for celula in aba_ativa[intervalo]:
            # Se o ativo atual da planilha estiver dentro da lista_ativos, ele analisa o ativo:
            if celula[0].value in lista_ativos:
                try:  # Caso o valor venha como uma string.
                    valor_celula: str = celula[posicao_do_elemento_valorizacao].value.replace(',',
                                                                                              '.')  # -> Alterando a ,
                    if '%' in valor_celula:
                        valor_celula: float = float(valor_celula.replace('%', ''))  # -> Alterando a %
                    else:
                        valor_celula: float = float(valor_celula) * 100  # -> Trazendo o valor de volta ao normal.
                except AttributeError:  # Caso o valor já venha como float.
                    valor_celula: float = celula[posicao_do_elemento_valorizacao].value
                    valor_celula: float = float(valor_celula) * 100
                # Se o valor da Valorização estiver acima do valor minimo aceitável: (verde)
                if valor_celula >= valor_minimo_aceitavel:
                    celula[posicao_do_elemento_valorizacao].fill = PatternFill(start_color='008000', end_color=None,
                                                                               fill_type='solid')
                # Se o valor da Valorização estiver entre -5 e 0: (amarelo)
                elif -5 <= valor_celula <= 0:
                    celula[posicao_do_elemento_valorizacao].fill = PatternFill(start_color='FFFF00', end_color=None,
                                                                               fill_type='solid')
                # Se o valor da Valorização estiver abaixo do valor minimo aceitável: (vermelho)
                else:
                    celula[posicao_do_elemento_valorizacao].fill = PatternFill(start_color='FFFF0000', end_color=None,
                                                                               fill_type='solid')

                lista_ativos.remove(celula[0].value)  # -> Remove o ativo que foi analisado da lista_ativos.
                if len(lista_ativos) == 0:  # -> Se não existem mais ativos para serem analisados, encerra o loop.
                    break
    except Exception as e:
        print(f'Error! - 4 - {e}')
    else:
        planilha.save('PlanilhaExcel\\Arquivo.xlsx')
    finally:
        planilha.close()


if __name__ == '__main__':
    # analisar_pvp_excel('fiis', ['RBRY11', 'VGIR11'])
    # analisar_pvp_excel('fiis', ['DEVA11', 'HGLG11'])
    analisar_pvp_excel('fiis', lista_ativos=None, todos=True)

    # analisar_pvp_excel('acoes', ['TAEE4', 'BBDC4'])
    # analisar_pvp_excel('acoes', ['JBSS3', 'EKTR3'])
    analisar_pvp_excel('acoes', lista_ativos=None, todos=True)
